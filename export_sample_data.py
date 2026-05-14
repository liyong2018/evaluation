"""
随机抽取四川省内5-6个区/县，导出该区县下所有乡镇和所有社区数据为Excel。
- 每个区/县导出其下所有乡镇数据(survey_data)和所有社区数据(community_disaster_reduction_capacity)
- 2024年和2025年各一套
- 排除开发区、高新区等
- 列名使用中文，与页面一致
"""
import os
import pymysql
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

DB_CONFIG = {
    'host': '127.0.0.1',
    'port': 30314,
    'user': 'root',
    'password': '123456',
    'database': 'evaluate_db',
    'charset': 'utf8mb4',
}

OUTPUT_DIR = '/Users/lql/Documents/data/Evaluation/evaluation/export_sample'

TOWNSHIP_COLUMNS = [
    ('区域名称', None),
    ('省份', 'province'),
    ('市/州', 'city'),
    ('区/县/市', 'county'),
    ('街道/乡镇', 'township'),
    ('人口数量', 'population'),
    ('管理人员', 'management_staff'),
    ('风险评估', 'risk_assessment'),
    ('资金投入(万元)', 'funding_amount'),
    ('物资价值(万元)', 'material_value'),
    ('医院床位', 'hospital_beds'),
    ('消防员数量', 'firefighters'),
    ('志愿者人数', 'volunteers'),
    ('民兵预备役', 'militia_reserve'),
    ('培训参与人次', 'training_participants'),
    ('避难场所容量', 'shelter_capacity'),
    ('创建时间', 'create_time'),
]

COMMUNITY_COLUMNS = [
    ('省份', 'province_name'),
    ('市/州', 'city_name'),
    ('区/县/市', 'county_name'),
    ('街道/乡镇', 'township_name'),
    ('社区(行政村)', 'community_name'),
    ('应急预案', 'has_emergency_plan'),
    ('弱势人群清单', 'has_vulnerable_groups_list'),
    ('地质灾害隐患点清单', 'has_disaster_points_list'),
    ('灾害类地图', 'has_disaster_map'),
    ('人口数量', 'resident_population'),
    ('资金投入(万元)', 'last_year_funding_amount'),
    ('物资价值(万元)', 'materials_equipment_value'),
    ('医疗服务点数', 'medical_service_count'),
    ('志愿者人数', 'registered_volunteer_count'),
    ('民兵预备役', 'militia_reserve_count'),
    ('培训参与人次', 'last_year_training_participants'),
    ('演练参与人次', 'last_year_drill_participants'),
    ('避难场所容量', 'emergency_shelter_capacity'),
    ('创建时间', 'create_time'),
]


def get_connection():
    return pymysql.connect(**DB_CONFIG)


def get_region_name(row):
    parts = [row.get('province', ''), row.get('city', ''),
             row.get('county', ''), row.get('township', '')]
    return ''.join(p for p in parts if p)


def sample_counties(cursor, year, count=6):
    """随机抽取指定年份的四川省区/县（排除开发区、高新区、天府新区等）"""
    cursor.execute(
        'SELECT DISTINCT county, city FROM survey_data '
        'WHERE province LIKE %s AND year = %s '
        'AND township NOT LIKE %s AND township NOT LIKE %s '
        'AND township NOT LIKE %s AND township NOT LIKE %s '
        'AND township NOT LIKE %s '
        'AND county NOT LIKE %s AND county NOT LIKE %s '
        'AND county NOT LIKE %s AND county NOT LIKE %s '
        'AND county NOT LIKE %s '
        'ORDER BY RAND() LIMIT %s',
        ('%四川%', year,
         '%开发%', '%高新%', '%经开%', '%工业园%', '%天府新区%',
         '%开发%', '%高新%', '%经开%', '%工业园%', '%天府新区%',
         count)
    )
    return cursor.fetchall()


def fetch_county_township_data(cursor, county, city, year):
    """获取指定区/县下所有乡镇的survey_data记录"""
    cursor.execute(
        'SELECT * FROM survey_data '
        'WHERE province LIKE %s AND year = %s '
        'AND city = %s AND county = %s '
        'AND township NOT LIKE %s AND township NOT LIKE %s '
        'AND township NOT LIKE %s AND township NOT LIKE %s '
        'AND township NOT LIKE %s '
        'ORDER BY township',
        ('%四川%', year, city, county, '%开发%', '%高新%', '%经开%', '%工业园%', '%天府新区%')
    )
    columns = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()
    return [dict(zip(columns, row)) for row in rows]


def fetch_county_community_data(cursor, county_name, city_name, year):
    """获取指定区/县下所有社区/行政村数据"""
    cursor.execute(
        'SELECT * FROM community_disaster_reduction_capacity '
        'WHERE province_name LIKE %s AND year = %s '
        'AND city_name = %s AND county_name = %s '
        'AND township_name NOT LIKE %s AND township_name NOT LIKE %s '
        'AND township_name NOT LIKE %s AND township_name NOT LIKE %s '
        'AND township_name NOT LIKE %s '
        'ORDER BY township_name, community_name',
        ('%四川%', year, city_name, county_name, '%开发%', '%高新%', '%经开%', '%工业园%', '%天府新区%')
    )
    columns = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()
    return [dict(zip(columns, row)) for row in rows]


def write_excel(filepath, headers, data_rows):
    """写入Excel文件，带格式"""
    wb = Workbook()
    ws = wb.active

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, size=11, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in data_rows:
            val = row[col_idx - 1]
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 30)

    wb.save(filepath)
    print(f'  已导出: {filepath} ({len(data_rows)} 条记录)')


def format_val(val):
    if val is None:
        return ''
    if isinstance(val, float):
        return round(val, 2)
    return val


def build_row(rec, columns):
    row = []
    for col_name, field in columns:
        if col_name == '区域名称':
            row.append(get_region_name(rec))
        elif field:
            row.append(format_val(rec.get(field)))
        else:
            row.append('')
    return row


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    conn = get_connection()
    cursor = conn.cursor()

    for year in [2024, 2025]:
        print(f'\n===== {year}年 =====')

        counties = sample_counties(cursor, year, count=6)
        print(f'  抽取到的区/县:')
        for county, city in counties:
            print(f'    {city} - {county}')

        # === 乡镇数据：每个区/县所有乡镇 ===
        township_headers = [h for h, _ in TOWNSHIP_COLUMNS]
        for county, city in counties:
            records = fetch_county_township_data(cursor, county, city, year)
            rows = [build_row(rec, TOWNSHIP_COLUMNS) for rec in records]
            filename = f'{year}年_{city}_{county}_乡镇数据.xlsx'
            write_excel(os.path.join(OUTPUT_DIR, filename), township_headers, rows)

        # === 社区数据：每个区/县所有社区/行政村 ===
        community_headers = [h for h, _ in COMMUNITY_COLUMNS]
        for county, city in counties:
            records = fetch_county_community_data(cursor, county, city, year)
            rows = [build_row(rec, COMMUNITY_COLUMNS) for rec in records]
            filename = f'{year}年_{city}_{county}_社区数据.xlsx'
            write_excel(os.path.join(OUTPUT_DIR, filename), community_headers, rows)

    cursor.close()
    conn.close()
    print(f'\n全部完成！文件保存在: {OUTPUT_DIR}')


if __name__ == '__main__':
    main()
