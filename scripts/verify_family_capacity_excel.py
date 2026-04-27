#!/usr/bin/env python3
"""按样例 Excel 口径重算家庭减灾能力，并校验结果是否一致。"""

from collections import OrderedDict
from pathlib import Path
import math

from openpyxl import load_workbook


WORKBOOK_PATH = Path(__file__).resolve().parents[1] / "docs" / "副本07样例数据_家庭减灾能力.xlsx"
COUNTY_CODES = ["510603", "510604", "510623", "510681", "510682", "510683"]
COMBINED_WEIGHTS = [0.132, 0.108, 0.0693, 0.0693, 0.0714, 0.0899, 0.0957, 0.1044, 0.0806, 0.065, 0.0572, 0.0572]
L1_GROUPS = [
    ((0, 1), [0.55, 0.45]),
    ((2, 3, 4), [0.33, 0.33, 0.34]),
    ((5, 6, 7), [0.31, 0.33, 0.36]),
    ((8, 9, 10, 11), [0.31, 0.25, 0.22, 0.22]),
]


def load_county_aggregates():
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb["指标赋值"]
    aggregates = OrderedDict()
    for row in range(2, ws.max_row + 1):
        region = str(ws.cell(row, 1).value)[:6]
        if region not in aggregates:
            aggregates[region] = [0.0] * 15
        values = [ws.cell(row, col).value for col in range(7, 22)]
        for idx, value in enumerate(values):
            aggregates[region][idx] += float(value or 0)
    return aggregates, wb


def vector_normalize(rows):
    cols = list(zip(*rows.values()))
    denominators = [math.sqrt(sum(x * x for x in col)) for col in cols]
    normalized = OrderedDict()
    for region, values in rows.items():
        normalized[region] = [values[i] / denominators[i] if denominators[i] else 0.0 for i in range(len(values))]
    return normalized


def topsis_scores(weighted_rows):
    maxima = [max(col) for col in zip(*weighted_rows.values())]
    minima = [min(col) for col in zip(*weighted_rows.values())]
    scores = OrderedDict()
    for region, values in weighted_rows.items():
        d_plus = math.sqrt(sum((maxima[i] - values[i]) ** 2 for i in range(len(values))))
        d_minus = math.sqrt(sum((minima[i] - values[i]) ** 2 for i in range(len(values))))
        ci = d_minus / (d_plus + d_minus) if (d_plus + d_minus) else 1.0
        scores[region] = (d_plus, d_minus, ci)
    return scores


def extract_excel_scores(wb, sheet_name, start_row, score_col, grade_col):
    ws = wb[sheet_name]
    result = OrderedDict()
    for idx, region in enumerate(COUNTY_CODES, start=start_row):
        result[region] = {
            "score": float(ws.cell(idx, score_col).value),
            "grade": ws.cell(idx, grade_col).value,
        }
    return result


def determine_grade(score, mean, stdev):
    if mean <= 0.5 * stdev:
        if score >= mean + 1.5 * stdev:
            return "强"
        if score >= mean + 0.5 * stdev:
            return "较强"
        return "中等"
    if mean <= 1.5 * stdev:
        if score >= mean + 1.5 * stdev:
            return "强"
        if score >= mean + 0.5 * stdev:
            return "较强"
        if score >= mean - 0.5 * stdev:
            return "中等"
        return "较弱"
    if score >= mean + 1.5 * stdev:
        return "强"
    if score >= mean + 0.5 * stdev:
        return "较强"
    if score >= mean - 0.5 * stdev:
        return "中等"
    if score >= mean - 1.5 * stdev:
        return "较弱"
    return "弱"


def main():
    county_aggregates, wb = load_county_aggregates()

    assigned = OrderedDict()
    for region, values in county_aggregates.items():
        total_people = values[3]
        assigned[region] = [
            (values[0] + values[1] + values[2]) / total_people if total_people else 0.0,
            values[4] / total_people if total_people else 0.0,
            values[5],
            values[6],
            values[7],
            values[8],
            values[9],
            values[10],
            values[11],
            values[12],
            values[13],
            values[14],
        ]

    normalized = vector_normalize(assigned)

    weighted = OrderedDict()
    for region in COUNTY_CODES:
        row = [assigned[region][0] * COMBINED_WEIGHTS[0]]
        row.extend(normalized[region][idx] * COMBINED_WEIGHTS[idx] for idx in range(1, 12))
        weighted[region] = row

    overall_scores = topsis_scores(weighted)
    excel_overall = extract_excel_scores(wb, "家庭减灾能力", 76, 5, 5)

    failures = []
    for region in COUNTY_CODES:
        expected = excel_overall[region]["score"]
        actual = overall_scores[region][2]
        if abs(expected - actual) > 1e-10:
            failures.append((region, "overall_score", actual, expected))

    overall_values = [overall_scores[region][2] for region in COUNTY_CODES]
    overall_mean = sum(overall_values) / len(overall_values)
    overall_stdev = math.sqrt(sum((value - overall_mean) ** 2 for value in overall_values) / (len(overall_values) - 1))
    for region in COUNTY_CODES:
        actual_grade = determine_grade(overall_scores[region][2], overall_mean, overall_stdev)
        excel_grade = wb["家庭减灾能力"][f"E{87 + COUNTY_CODES.index(region)}"].value
        if actual_grade != excel_grade:
            failures.append((region, "overall_grade", actual_grade, excel_grade))

    for group_index, (group_cols, group_weights) in enumerate(L1_GROUPS, start=1):
        weighted_rows = OrderedDict()
        for region in COUNTY_CODES:
            row = []
            for offset, col in enumerate(group_cols):
                if group_index == 1 and offset == 0:
                    row.append(assigned[region][col] * group_weights[offset])
                else:
                    row.append(normalized[region][col] * group_weights[offset])
            weighted_rows[region] = row

        group_scores = topsis_scores(weighted_rows)
        score_col = 4 + group_index + 1
        grade_col = score_col
        excel_sheet = wb["一级指标"]
        score_row_base = 76
        grade_row_base = 87
        values = [group_scores[region][2] for region in COUNTY_CODES]
        mean = sum(values) / len(values)
        stdev = math.sqrt(sum((value - mean) ** 2 for value in values) / (len(values) - 1))
        for idx, region in enumerate(COUNTY_CODES):
            excel_score = float(excel_sheet.cell(score_row_base + idx, 4 + group_index).value)
            actual_score = group_scores[region][2]
            if abs(excel_score - actual_score) > 1e-10:
                failures.append((region, f"l1_{group_index}_score", actual_score, excel_score))
            excel_grade = excel_sheet.cell(grade_row_base + idx, 4 + group_index).value
            actual_grade = determine_grade(actual_score, mean, stdev)
            if actual_grade != excel_grade:
                failures.append((region, f"l1_{group_index}_grade", actual_grade, excel_grade))

    if failures:
        print("FAILED")
        for failure in failures:
            print(failure)
        raise SystemExit(1)

    print("PASS: Excel家庭减灾能力口径复算一致")
    for region in COUNTY_CODES:
        print(region, round(overall_scores[region][2], 12))


if __name__ == "__main__":
    main()
