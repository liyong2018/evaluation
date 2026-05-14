#!/usr/bin/env python3
"""
Export township and community survey data for sampled counties.

Inputs:
- export_sample/2024年组织机构_按市州抽取30%区县_排除功能区.xlsx
- export_sample/2025年组织机构_按市州抽取30%区县_排除功能区.xlsx

The output workbook shape follows the existing 新津区 sample files.
"""

from __future__ import annotations

import argparse
import copy
import datetime as dt
import hashlib
import json
import math
import re
import shutil
import zipfile
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence, Tuple

import openpyxl
import pymysql


ROOT = Path(__file__).resolve().parents[1]
EXPORT_SAMPLE = ROOT / "export_sample"
TOWNSHIP_SAMPLE = EXPORT_SAMPLE / "2024年_成都市_新津区_乡镇数据.xlsx"
COMMUNITY_SAMPLE = EXPORT_SAMPLE / "2024年_成都市_新津区_社区数据.xlsx"

TOWNSHIP_HEADERS = [
    "区域名称",
    "省份",
    "市/州",
    "区/县/市",
    "街道/乡镇",
    "人口数量",
    "管理人员",
    "风险评估",
    "资金投入(万元)",
    "物资价值(万元)",
    "医院床位",
    "消防员数量",
    "志愿者人数",
    "民兵预备役",
    "培训参与人次",
    "避难场所容量",
    "创建时间",
]

COMMUNITY_HEADERS = [
    "省份",
    "市/州",
    "区/县/市",
    "街道/乡镇",
    "社区(行政村)",
    "应急预案",
    "弱势人群清单",
    "地质灾害隐患点清单",
    "灾害类地图",
    "人口数量",
    "资金投入(万元)",
    "物资价值(万元)",
    "医疗服务点数",
    "志愿者人数",
    "民兵预备役",
    "培训参与人次",
    "演练参与人次",
    "避难场所容量",
    "创建时间",
]

FUNCTIONAL_COUNTY_KEYWORDS = [
    "新区",
    "高新区",
    "开发区",
    "技术产业开发区",
    "产业开发区",
    "经济开发区",
    "经开区",
    "商贸园",
    "经济区",
    "风景区",
    "示范区",
    "园区",
    "科学城",
    "直管区",
]


def yes_no(value: Any) -> Any:
    if value is None:
        return None
    text = str(value).strip()
    if text in {"1", "是", "true", "TRUE", "True", "Y", "yes"}:
        return "是"
    if text in {"0", "否", "false", "FALSE", "False", "N", "no"}:
        return "否"
    return value


def safe_filename_part(value: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', "_", str(value).strip())


def normalize_datetime(value: Any) -> Any:
    if isinstance(value, (dt.datetime, dt.date)):
        return value
    return value


def copy_sheet_style(template_path: Path, headers: Sequence[str]) -> openpyxl.Workbook:
    src = openpyxl.load_workbook(template_path)
    src_ws = src.active
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = src_ws.title

    for col_idx, header in enumerate(headers, 1):
        src_cell = src_ws.cell(row=1, column=col_idx)
        cell = ws.cell(row=1, column=col_idx, value=header)
        if src_cell.has_style:
            cell.font = copy.copy(src_cell.font)
            cell.fill = copy.copy(src_cell.fill)
            cell.border = copy.copy(src_cell.border)
            cell.alignment = copy.copy(src_cell.alignment)
            cell.number_format = src_cell.number_format
            cell.protection = copy.copy(src_cell.protection)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = (
            src_ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width or 16
        )
    if src_ws.freeze_panes:
        ws.freeze_panes = src_ws.freeze_panes
    else:
        ws.freeze_panes = "A2"
    return wb


def style_data_rows(ws: openpyxl.worksheet.worksheet.Worksheet, template_path: Path, row_count: int, col_count: int) -> None:
    src = openpyxl.load_workbook(template_path)
    src_ws = src.active
    template_row = 2 if src_ws.max_row >= 2 else 1
    for row_idx in range(2, row_count + 2):
        for col_idx in range(1, col_count + 1):
            src_cell = src_ws.cell(row=template_row, column=col_idx)
            cell = ws.cell(row=row_idx, column=col_idx)
            if src_cell.has_style:
                cell.font = copy.copy(src_cell.font)
                cell.fill = copy.copy(src_cell.fill)
                cell.border = copy.copy(src_cell.border)
                cell.alignment = copy.copy(src_cell.alignment)
                cell.number_format = src_cell.number_format
                cell.protection = copy.copy(src_cell.protection)
    if row_count > 0:
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(col_count)}{row_count + 1}"


def read_counties(sample_path: Path) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(sample_path, read_only=True, data_only=True)
    ws = wb["抽样区县清单"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(v).strip() for v in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}
    counties = []
    for row in rows[1:]:
        if not row or row[idx["区县代码"]] is None:
            continue
        counties.append(
            {
                "year": int(row[idx["年份"]]),
                "city": str(row[idx["市州"]]).strip(),
                "county_code": str(row[idx["区县代码"]]).strip(),
                "county": str(row[idx["区县名称"]]).strip(),
                "sample_order": int(row[idx["抽样序号"]]),
            }
        )
    return counties


def is_functional_county_name(name: str) -> bool:
    return any(keyword in (name or "") for keyword in FUNCTIONAL_COUNTY_KEYWORDS)


def fetch_effective_counties(conn, target_year: int, exclude_functional: bool = True) -> List[Dict[str, Any]]:
    sql = """
        SELECT code, name, city_name, is_deleted, year, id
        FROM organization
        WHERE level = 3
          AND year <= %s
          AND city_name IS NOT NULL
          AND city_name <> ''
        ORDER BY code ASC, year DESC, id DESC
    """
    with conn.cursor() as cur:
        cur.execute(sql, (target_year,))
        rows = cur.fetchall()

    latest_by_code: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        code = str(row["code"]).strip()
        if code not in latest_by_code:
            latest_by_code[code] = row

    counties = []
    for row in latest_by_code.values():
        if int(row.get("is_deleted") or 0) != 0:
            continue
        name = str(row["name"]).strip()
        if exclude_functional and is_functional_county_name(name):
            continue
        counties.append(
            {
                "year": target_year,
                "city": str(row["city_name"]).strip(),
                "county_code": str(row["code"]).strip(),
                "county": name,
            }
        )
    counties.sort(key=lambda item: (item["city"], item["county_code"]))
    return counties


def sample_counties_by_city(
    counties: List[Dict[str, Any]],
    sample_ratio: float,
    seed: str,
) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for county in counties:
        grouped.setdefault(county["city"], []).append(county)

    sampled: List[Dict[str, Any]] = []
    for city in sorted(grouped):
        city_counties = grouped[city]
        sample_count = max(1, int(math.ceil(len(city_counties) * sample_ratio)))

        def sort_key(item: Dict[str, Any]) -> Tuple[str, str]:
            text = f"{seed}|{city}|{item['county_code']}"
            return hashlib.sha256(text.encode("utf-8")).hexdigest(), item["county_code"]

        for rank, county in enumerate(sorted(city_counties, key=sort_key)[:sample_count], 1):
            sampled.append(
                {
                    **county,
                    "city_county_count": len(city_counties),
                    "sample_count": sample_count,
                    "sample_order": rank,
                }
            )
    sampled.sort(key=lambda item: (item["city"], item["sample_order"], item["county_code"]))
    return sampled


def fetch_organization_detail_rows(conn, counties: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    detail_rows: List[Dict[str, Any]] = []
    for county in counties:
        year = county["year"]
        county_code = county["county_code"]
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id
                FROM organization
                WHERE code = %s AND year <= %s AND level = 3 AND is_deleted = 0
                ORDER BY year DESC, id DESC
                LIMIT 1
                """,
                (county_code, year),
            )
            county_org = cur.fetchone()
        county_id = county_org["id"] if county_org else None
        if county_id is None:
            continue
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT code, name, level, township_name, community_name
                FROM grassroots_organization
                WHERE county_id = %s
                  AND year <= %s
                  AND is_deleted = 0
                ORDER BY level ASC, code ASC, year DESC, id DESC
                """,
                (county_id, year),
            )
            rows = cur.fetchall()

        seen = set()
        for row in rows:
            code = str(row["code"]).strip()
            level = int(row["level"])
            key = (code, level)
            if key in seen:
                continue
            seen.add(key)
            if level == 4:
                township_code = code
                township_name = row.get("township_name") or row.get("name")
                community_code = None
                community_name = None
                level_name = "乡镇级"
            elif level == 5:
                township_code = code[:9] if len(code) >= 9 else None
                township_name = row.get("township_name")
                community_code = code
                community_name = row.get("community_name") or row.get("name")
                level_name = "社区级"
            else:
                continue
            detail_rows.append(
                {
                    "年份": year,
                    "市州": county["city"],
                    "区县代码": county_code,
                    "区县名称": county["county"],
                    "层级": level_name,
                    "组织机构代码": code,
                    "组织机构名称": row.get("name"),
                    "乡镇代码": township_code,
                    "乡镇名称": township_name,
                    "社区代码": community_code,
                    "社区名称": community_name,
                }
            )
    return detail_rows


def create_sample_organization_workbook(path: Path, counties: List[Dict[str, Any]], detail_rows: List[Dict[str, Any]]) -> None:
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "市州抽样汇总"
    ws_counties = wb.create_sheet("抽样区县清单")
    ws_detail = wb.create_sheet("组织机构明细")

    summary_headers = ["年份", "市州", "区县总数", "抽样区县数", "抽样比例", "抽中区县"]
    ws_summary.append(summary_headers)
    by_city: Dict[str, List[Dict[str, Any]]] = {}
    for county in counties:
        by_city.setdefault(county["city"], []).append(county)
    for city in sorted(by_city):
        rows = sorted(by_city[city], key=lambda item: item["sample_order"])
        total = rows[0]["city_county_count"]
        sample_count = rows[0]["sample_count"]
        ws_summary.append([
            rows[0]["year"],
            city,
            total,
            sample_count,
            sample_count / total if total else None,
            "、".join(row["county"] for row in rows),
        ])

    county_headers = ["年份", "市州", "区县代码", "区县名称", "市州区县总数", "抽样区县数", "抽样序号"]
    ws_counties.append(county_headers)
    for county in counties:
        ws_counties.append([
            county["year"],
            county["city"],
            county["county_code"],
            county["county"],
            county["city_county_count"],
            county["sample_count"],
            county["sample_order"],
        ])

    detail_headers = ["年份", "市州", "区县代码", "区县名称", "层级", "组织机构代码", "组织机构名称", "乡镇代码", "乡镇名称", "社区代码", "社区名称"]
    ws_detail.append(detail_headers)
    for row in detail_rows:
        ws_detail.append([row.get(header) for header in detail_headers])

    for ws in [ws_summary, ws_counties, ws_detail]:
        for col_idx, cell in enumerate(ws[1], 1):
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="4F81BD")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(36, max(12, len(str(cell.value)) * 2 + 4))
        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def connect(args: argparse.Namespace):
    return pymysql.connect(
        host=args.host,
        port=args.port,
        user=args.user,
        password=args.password,
        database=args.database,
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
    )


def fetch_township_rows(conn, year: int, county_code: str) -> List[Dict[str, Any]]:
    sql = """
        SELECT
            COALESCE(township_address, CONCAT(COALESCE(province, ''), COALESCE(city, ''), COALESCE(county, ''), COALESCE(township, ''))) AS region_name,
            province,
            city,
            county,
            township,
            population,
            management_staff,
            risk_assessment,
            funding_amount,
            material_value,
            hospital_beds,
            firefighters AS firefighters_count,
            volunteers AS volunteers_count,
            militia_reserve AS militia_reserve_count,
            training_participants,
            shelter_capacity,
            create_time,
            region_code
        FROM survey_data
        WHERE is_deleted = 0
          AND year = %s
          AND region_code LIKE %s
        ORDER BY region_code ASC, create_time DESC
    """
    with conn.cursor() as cur:
        cur.execute(sql, (year, county_code + "%"))
        rows = cur.fetchall()
    return rows


def fetch_community_rows(conn, year: int, county_code: str) -> List[Dict[str, Any]]:
    sql = """
        SELECT
            province_name,
            city_name,
            county_name,
            township_name,
            community_name,
            has_emergency_plan,
            has_vulnerable_groups_list,
            has_disaster_points_list,
            has_disaster_map,
            resident_population,
            last_year_funding_amount,
            materials_equipment_value,
            medical_service_count,
            registered_volunteer_count,
            militia_reserve_count,
            last_year_training_participants,
            last_year_drill_participants,
            emergency_shelter_capacity,
            create_time,
            region_code
        FROM community_disaster_reduction_capacity
        WHERE year = %s
          AND region_code LIKE %s
        ORDER BY region_code ASC, create_time DESC
    """
    with conn.cursor() as cur:
        cur.execute(sql, (year, county_code + "%"))
        rows = cur.fetchall()
    return rows


def write_township_workbook(path: Path, rows: List[Dict[str, Any]]) -> None:
    wb = copy_sheet_style(TOWNSHIP_SAMPLE, TOWNSHIP_HEADERS)
    ws = wb.active
    for r_idx, row in enumerate(rows, 2):
        values = [
            row.get("region_name"),
            row.get("province"),
            row.get("city"),
            row.get("county"),
            row.get("township"),
            row.get("population"),
            row.get("management_staff"),
            yes_no(row.get("risk_assessment")),
            row.get("funding_amount"),
            row.get("material_value"),
            row.get("hospital_beds"),
            row.get("firefighters_count"),
            row.get("volunteers_count"),
            row.get("militia_reserve_count"),
            row.get("training_participants"),
            row.get("shelter_capacity"),
            normalize_datetime(row.get("create_time")),
        ]
        for c_idx, value in enumerate(values, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    style_data_rows(ws, TOWNSHIP_SAMPLE, len(rows), len(TOWNSHIP_HEADERS))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_community_workbook(path: Path, rows: List[Dict[str, Any]]) -> None:
    wb = copy_sheet_style(COMMUNITY_SAMPLE, COMMUNITY_HEADERS)
    ws = wb.active
    for r_idx, row in enumerate(rows, 2):
        values = [
            row.get("province_name"),
            row.get("city_name"),
            row.get("county_name"),
            row.get("township_name"),
            row.get("community_name"),
            yes_no(row.get("has_emergency_plan")),
            yes_no(row.get("has_vulnerable_groups_list")),
            yes_no(row.get("has_disaster_points_list")),
            yes_no(row.get("has_disaster_map")),
            row.get("resident_population"),
            row.get("last_year_funding_amount"),
            row.get("materials_equipment_value"),
            row.get("medical_service_count"),
            row.get("registered_volunteer_count"),
            row.get("militia_reserve_count"),
            row.get("last_year_training_participants"),
            row.get("last_year_drill_participants"),
            row.get("emergency_shelter_capacity"),
            normalize_datetime(row.get("create_time")),
        ]
        for c_idx, value in enumerate(values, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    style_data_rows(ws, COMMUNITY_SAMPLE, len(rows), len(COMMUNITY_HEADERS))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def create_manifest(path: Path, records: List[Dict[str, Any]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "导出清单"
    headers = ["年份", "市州", "区县代码", "区县名称", "乡镇数据条数", "社区数据条数", "乡镇文件", "社区文件"]
    ws.append(headers)
    for record in records:
        ws.append([
            record["year"],
            record["city"],
            record["county_code"],
            record["county"],
            record["township_count"],
            record["community_count"],
            record["township_file"],
            record["community_file"],
        ])
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="4F81BD")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(12, min(48, len(header) * 2 + 4))
    ws.auto_filter.ref = f"A1:H{len(records) + 1}"
    ws.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def zip_outputs(zip_path: Path, files: Iterable[Path], base_dir: Path) -> None:
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for file_path in files:
            zf.write(file_path, file_path.relative_to(base_dir))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=30314)
    parser.add_argument("--user", default="root")
    parser.add_argument("--password", default="123456")
    parser.add_argument("--database", default="evaluate_db")
    parser.add_argument("--output-dir", default=str(EXPORT_SAMPLE / "sample_county_data_exports"))
    parser.add_argument("--sample-year", type=int, default=None)
    parser.add_argument("--sample-ratio", type=float, default=0.30)
    parser.add_argument("--sample-seed", default="county-sample-30pct-20260506")
    parser.add_argument("--include-functional-counties", action="store_true")
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    if output_dir.exists():
        shutil.rmtree(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    counties: List[Dict[str, Any]] = []
    records: List[Dict[str, Any]] = []
    generated_files: List[Path] = []
    conn = connect(args)
    try:
        if args.sample_year is not None:
            effective_counties = fetch_effective_counties(
                conn,
                args.sample_year,
                exclude_functional=not args.include_functional_counties,
            )
            counties = sample_counties_by_city(effective_counties, args.sample_ratio, args.sample_seed)
            detail_rows = fetch_organization_detail_rows(conn, counties)
            sample_workbook_path = EXPORT_SAMPLE / f"{args.sample_year}年组织机构_按市州抽取30%区县_排除功能区.xlsx"
            create_sample_organization_workbook(sample_workbook_path, counties, detail_rows)
            sample_workbook_copy = output_dir / sample_workbook_path.name
            shutil.copy2(sample_workbook_path, sample_workbook_copy)
            generated_files.append(sample_workbook_copy)
        else:
            sample_files = [
                EXPORT_SAMPLE / "2024年组织机构_按市州抽取30%区县_排除功能区.xlsx",
                EXPORT_SAMPLE / "2025年组织机构_按市州抽取30%区县_排除功能区.xlsx",
            ]
            for sample_file in sample_files:
                counties.extend(read_counties(sample_file))

        for county in counties:
            year = county["year"]
            city = county["city"]
            county_name = county["county"]
            county_code = county["county_code"]
            prefix = f"{year}年_{safe_filename_part(city)}_{safe_filename_part(county_name)}"

            township_rows = fetch_township_rows(conn, year, county_code)
            community_rows = fetch_community_rows(conn, year, county_code)

            township_path = output_dir / f"{prefix}_乡镇数据.xlsx"
            community_path = output_dir / f"{prefix}_社区数据.xlsx"
            write_township_workbook(township_path, township_rows)
            write_community_workbook(community_path, community_rows)
            generated_files.extend([township_path, community_path])

            records.append(
                {
                    **county,
                    "township_count": len(township_rows),
                    "community_count": len(community_rows),
                    "township_file": township_path.name,
                    "community_file": community_path.name,
                }
            )
            print(f"{year} {city} {county_code} {county_name}: township={len(township_rows)} community={len(community_rows)}")
    finally:
        conn.close()

    manifest_path = output_dir / "导出清单.xlsx"
    create_manifest(manifest_path, records)
    generated_files.append(manifest_path)

    summary_path = output_dir / "导出统计.json"
    summary = {
        "county_count": len(records),
        "workbook_count": len([p for p in generated_files if p.name.endswith("_乡镇数据.xlsx") or p.name.endswith("_社区数据.xlsx")]),
        "total_township_rows": sum(r["township_count"] for r in records),
        "total_community_rows": sum(r["community_count"] for r in records),
        "zero_township_counties": [r for r in records if r["township_count"] == 0],
        "zero_community_counties": [r for r in records if r["community_count"] == 0],
    }
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    zip_path = output_dir.parent / f"{output_dir.name}.zip"
    if zip_path.exists():
        zip_path.unlink()
    zip_outputs(zip_path, generated_files + [summary_path], output_dir)
    print("SUMMARY", json.dumps(summary, ensure_ascii=False))
    print("OUTPUT_DIR", output_dir)
    print("ZIP", zip_path)


if __name__ == "__main__":
    main()
